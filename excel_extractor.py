import os
import subprocess
from openpyxl import load_workbook
import shutil
import pathlib
from scripts.extractor.minio_tools import save_in_minio
from scripts.extractor.tools import extract_tags


def extract_image(excel_file):
    image_extensions = ("*.png", "*.jpg", "*.jpeg", "*.gif", "*.bmp", "*.tiff", "*.tif")
    if type(excel_file) is str:
        excel_file = pathlib.Path(excel_file)
    # temp_file = copy.deepcopy(excel_file)
    name = excel_file.name.replace("".join(excel_file.suffixes), "").replace(
        " ", ""
    )  # name of excel file without suffixes
    temp_file = pathlib.Path(excel_file).parent / "temp.xlsx"  # temp xlsx
    temp_zip = temp_file.with_suffix(".zip")  # temp zip
    shutil.copyfile(excel_file, temp_file)
    temp_file.rename(str(temp_zip))
    extract_dir = temp_file.parent / "temp"
    extract_dir.mkdir(exist_ok=True)
    shutil.unpack_archive(temp_zip, extract_dir)  # unzip xlsx zip file
    paths_img = []
    for ext in image_extensions:
        paths_img.extend(sorted((extract_dir / "xl" / "media").glob(ext)))
    # sorted((extract_dir / "xl" / "media").glob())  # find images
    move_paths = {
        path: pathlib.Path(path).parent / (name + f"-{str(n)}{path.suffix}")
        for n, path in enumerate(paths_img)
    }  # create move path dict
    new_paths = [
        shutil.move(old, new) for old, new in move_paths.items()
    ]  # move / rename image files

    images = []
    for loc in new_paths:
        image = save_in_minio(location=loc)
        if image:
            images.append(image)
    shutil.rmtree(extract_dir) # delete temp folder
    temp_zip.unlink() # delete temp file

    return images


def xls_to_xlsx(xls_filename: str, output_dir="tmp"):
    """Converts an XLS file to XLSX using xlrd and openpyxl."""
    subprocess.run(
        [
            "libreoffice",
            "--headless",
            "--convert-to",
            "xlsx",
            os.path.join(output_dir, xls_filename),
            "--outdir",
            output_dir,
        ],
        check=True,
    )

    # Cleanup the temporary docx file
    return os.path.join(output_dir, xls_filename.replace(".xls", ".xlsx"))


def extract_excel_data(excel_file):
    """
    Extracts data and images from an Excel file (xls or xlsx).

    Args:
      excel_file: Path to the Excel file.
    """
    if not (excel_file.endswith(".xls") or excel_file.endswith(".xlsx")):
        raise ValueError("path must be an excel file")

    if excel_file.endswith(".xls"):
        excel_file = xls_to_xlsx(excel_file)

    wb = load_workbook(excel_file)
    pages = []
    tags = []
    for page_num, sheet_name in enumerate(wb.sheetnames):
        page_data = {"id": page_num, "fulltext": ""}

        sheet = wb[sheet_name]
        # print(f"Sheet: {sheet_name}")

        # Extract data (example)
        rows = []
        for row in sheet.iter_rows():
            rows.append(" ".join([str(cell.value) for cell in row if cell.value]))

        page_data.update({"fulltext": "\n".join(rows)})
        pages.append(page_data)
        tags.extend(extract_tags(page_data.get("fulltext")))
        # Extract images
        # calling the image_loader
        
        

    images = extract_image(excel_file)
    path = pathlib.Path(excel_file)
    output_dict = {
        "name": path.parts[-1],
        "path": str(path),
        # "type": "TYPE",
        # "rubrique": "RU",
        "extension": path.name.split(".")[-1],
        "tags": tags,
        "tables": [],
        "images": images,
        "pages": pages,
    }

    return output_dict
